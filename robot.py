##################################################################################################################
# Use: python robot.py exchange simulation_flag(s/r/sns/rns) basic_currency altcoin entry_price TP% SL% [limit_of_amount_to sell] [sell_portion]
# s - simulation with stop-loss
# r - real mode with stop-loss 
# sns - simulation without stop loss and only with trailing stop on profit 
# rns - real mode without stop loss and only with trailing stop on profit 
#
# Example: trade 100 LTC (vs BTC) bought at 0.0017 with the target profit of 18% and stop-loss 5% in simulation mode
# > python robot.py btrx s BTC LTC 0.0017 18 5 100
#
# Running without input parameters will start Telegram bot listener
#
# Conservative - TP 1.13 / SL 0.95, volatile (NEO etc.) TP 1.18 / SL 0.94
# SL threshold applies both to original stop and to trailing stop

################################ Libraries ############################################
# Standard libraries 
import os
import time
import sys
from sys import exit, argv
from time import localtime, strftime
import subprocess   
import math
import smtplib
from email.MIMEMultipart import MIMEMultipart
from email.MIMEText import MIMEText
import urllib2
import decimal
from decimal import Decimal
from openpyxl import Workbook, load_workbook   
from openpyxl.styles import Font, Fill
import json # requests
from shutil import copyfile # to copy files
import numpy as np

# Decimal precision and roubding 
decimal.getcontext().prec = 25
decimal.getcontext().rounding = 'ROUND_DOWN'

## Custom libraries 
from telegramlib import telegram # my lib to work with Telegram
from sqltools import query_lastrow_id, query # proper requests to sqlite db
from loglib import logfile # logging 
import platformlib as platform  # detecting the OS and assigning proper folders 

# Universal functions for all exchanges              
from exchange_func import getticker, getopenorders, cancel, getorderhistory, getorder, getbalance, selllimit, getorderbook, buylimit, getbalances, binance_price_precise, binance_quantity_precise, getpositions, closepositions

################################ Config - part I ############################################

### TD analysis library
import tdlib as tdlib
td_info = tdlib.tdlib()

### Platform
platform = platform.platformlib()
platform_run, cmd_init, cmd_init_buy = platform.initialise() 
print "Initialising..."

### Set up the speedrun multiplier if need to test with higher speeds. 1 is normal, 2 is 2x faster 
speedrun = 1  

### Telegram integration  
chat = telegram()

comm_method = 'chat' # 'mail' or 'chat'
send_messages = True

### Command prompt parameters  

### Default values
no_input = False 
trailing_stop_flag = True  # default mode is to have trailing stop

### Input parameters 
try: 
    simulation_param = argv[1]
    if simulation_param == 's': 
        simulation = True
        stop_loss = True
    elif simulation_param == 'r': 
        simulation = False
        stop_loss = True
    elif simulation_param == 'sns': 
        simulation = True
        stop_loss = False
    elif simulation_param == 'rns': 
        simulation = False
        stop_loss = False
    elif simulation_param == 'rnts':
        trailing_stop_flag = False 
        simulation = False
        stop_loss = True
    else: 
        no_input = True 

    exchange_abbr = argv[2].lower()
    if exchange_abbr not in ['btrx', 'bina', 'bmex']: 
        print 'Incorrect exchange specified (should be btrx, bina, or bmex)\n\n'
        exit(0)
    if exchange_abbr == 'btrx': 
        exchange = 'bittrex' 
        comission_rate = 0.003
    elif exchange_abbr == 'bina': 
        exchange = 'binance' 
        comission_rate = 0.001
    elif exchange_abbr == 'bmex': 
        exchange = 'bitmex' 
        comission_rate = 0
        
    trade = argv[3].upper() 
    currency = argv[4].upper()
    price_curr = float(argv[5])

    price_target = float(argv[6])
    sl_target = float(argv[7])
    price_entry = price_curr

    tp = round(price_target/price_curr, 5)
    sl = round(sl_target/price_curr, 5) 
    tp_p = (tp - 1.0)*100.0 
    sl_p = (1.0 - sl)*100.0 

    try:
        limit_sell_amount = float(argv[8])
    except: 
        limit_sell_amount = 0
    try:
        sell_portion = float(argv[9])
    except: 
        sell_portion = None    
    # print 'Trade', trade, 'currency', currency, 'simulation', simulation, 'price_curr', price_curr, 'tp', tp, 'sl', sl, limit_sell_amount, sell_portion  #DEBUG   
except:
    no_input = True 

# Terminate if there is no proper input 
if no_input:
    print '----------------------------------------------------------------------------------------------\n' + \
    'Run parameters not specified. Restart the script using:\n' + \
    'robot.py simulation (s/r/sns/rns) exchange basic_curr altcoin entry_price TP SL [limit_of_amount_to_sell] [sell_portion]\n' +\
    'Example: > python robot.py s btrx BTC LTC 0.0017 0.0021 0.0015 100\n\n' +\
    'Modes:\n>s (simulation with stop-loss)\n>r (real mode with stop-loss)\n>sns (simulation and stop only on profit)\n>rns (real and stop only on profit)'  
    exit(0) 
    
###  If simulation and parameters are not specified 
if simulation is True:
    if limit_sell_amount == 0: 
        limit_sell_amount = 100
    simulation_balance = limit_sell_amount
    sell_portion = limit_sell_amount
    
### Market to trade 
market = '{0}-{1}'.format(trade, currency)
        
### Prices 
# something should be bought at a price_curr level to start from  

price_target = price_curr*tp
sl_target = price_curr*sl
price_entry = price_curr

#### Gmail login and pass (if used) 
fromaddr = "fromaddress@gmail.com"    # replace to a proper address 
toaddr = "to@address.com"    # replace to a proper address 
email_passw = "your_gmail_pass"

################################ Config - part II ############################################
### Intervals and timers in seconds  

sleep_timer = 30                 # Generic sleep timer. Applicable for the main monitoring loop and for the mooning procedure.
sleep_sale = 30                  # Sleep timer for sell orders to be filled 
flash_crash_ind = 0.5         # If something falls so much too fast - it is unusual and we should not sell (checking for 50% crashes)

## Interval and number of checks to get current (last) prices 
steps_ticker = 3 
sleep_ticker = 10               # so that ticker in total takes 30 seconds 

## Steps and timer for buybacks 
candle_steps = 80               # 100 for 5 min, 80 for 4
candle_sleep = 2.8              # Tested, 3 sec lead to having ~5 min 30 sec in between 

sleep_timer = int(sleep_timer/speedrun)
sleep_sale = int(sleep_sale/speedrun)
sleep_ticker = int(sleep_ticker/speedrun)
candle_steps = int(candle_steps/speedrun)

### To cancel buyback if there is an error and there were no sales made 
cancel_buyback = False 

### Starting variables  
main_curr_from_sell = 0     
commission_total = 0        
alt_sold_total = 0  
decrease_attempts_total = 0  
value_original = 0
stopped_mode = '' 

# Logger
logger = logfile(market, 'trade')

################################ Functions ############################################

### Log and print 
def lprint(arr):
    msg = ' '.join(map(lambda x: ''+ str(x), arr))
    logger.write(msg)
    print msg

    
##############################################  
##            Core get price / moon / sell functions           ##
##############################################

##################### Price comparison  
def strictly_increasing(L):
    return all(x<y for x, y in zip(L, L[1:]))
   
def equal_or_increasing(L):
    return all(x<=y for x, y in zip(L, L[1:]))

def strictly_decreasing(L):
    return all(x>y for x, y in zip(L, L[1:]))
    
def equal_or_decreasing(L):
    # Actually >= for our purposes
    return all(x>=y for x, y in zip(L, L[1:]))

##################### Processing sell outcome results and generating messages 
def process_stat(status): 
    global market
    global db, cur, job_id
    global cancel_buyback 
    
    flag = True   # default flag returned
    
    if status == 'stop':
        message = 'Finishing up normally'
        flag = False
        sql_string = "UPDATE jobs SET selling = 0 WHERE job_id = {}".format(job_id)     # DB update
        rows = query(sql_string)

    if status == 'err_low': 
        message = 'Trade amount was too small and returned error, finishing up'
        send_notification('Error: Too small trade', 'Too small trade to perform, finishing up')
        cancel_orders(market)
        flag = False
        cancel_buyback = True 
        
    if status == 'no_idea': 
        message = 'Sell calls did not return proper answer, aborting'
        send_notification('Error: No response from sell calls', 'Sell calls did not return proper answer, aborting')
        cancel_orders(market)
        flag = False
        cancel_buyback = True 
        
    if status == 'abort_telegram': 
        message = 'Aborted as requested via Telegram'
        cancel_orders(market)
        flag = False
        cancel_buyback = True 
        
    return flag, message

##################### Getting several last prices in short intervals instead of just one 
def get_last_price(market): 
    ticker_upd = {}
    price_upd = 0
    failed_attempts = 0
    for i in range(1, steps_ticker + 1):
        try:
            ticker_upd = getticker(exchange, market) 
            price_upd += ticker_upd
        except:
            # print "Issues with URL (!) for market", market
            failed_attempts += 1
        time.sleep(sleep_ticker)
        
    # Logging failed attempts number
    if failed_attempts > 0: 
        lprint(["Failed attempts to receive price:", failed_attempts])    
        
    # If retreiving prices fails completely
    if failed_attempts == steps_ticker:     
        ticker_upd = None  
        try:
            send_notification('Maintenance', market + ' seems to be on an automatic maintenance. Will try every 5 minutes.')
        except: 
            lprint(["Failed to send notification"])    
        while ticker_upd is None: 
            time.sleep(300) # sleeping for 5 minutes and checking again
            lprint(["Market could be on maintenance. Sleeping for 5 minutes."])    
            try:
                ticker_upd = getticker(exchange, market) 
            except: 
                ticker_upd = None
            price_upd = ticker_upd
    # If it is fine - get the average price 
    else: 
        price_upd = float(price_upd)/float(steps_ticker - failed_attempts)
        
    return price_upd

##################### Extreme in time series; returns value with the lowest or the highest ticker price among N-min intervals (candles) 
# type should be 'H' or 'L' (highest ore lowest in the series) 
def candle_extreme(type): 
    global market, candle_steps, candle_sleep
    ticker_upd = {}
    price_upd = 0
    price_extreme = 0
    failed_attempts = 0
    
    for i in range(1, candle_steps + 1): # 5 min: 100 checks x 3 sec (better indication than 30 checks x 10 sec); 80 x 3 for 4 minutes 
        try:
            ticker_upd = getticker(exchange, market) 
            price_upd = ticker_upd
            if type == 'L': 
                if (price_extreme == 0) or (price_upd < price_extreme): 
                    price_extreme = price_upd
            if type == 'H': 
                if (price_extreme == 0) or (price_upd > price_extreme): 
                    price_extreme = price_upd
        except:
            #print "Issues with URL (!) for market", market
            failed_attempts += 1
        time.sleep(candle_sleep) 
        
    # Logging failed attempts number
    if failed_attempts > 0: 
        lprint(["Failed attempts to receive price:", failed_attempts])    
        
    # If retreiving prices fails completely
    if failed_attempts == steps_ticker:     
        ticker_upd = None # Change
        # Could be related to maintenance
        try:
            send_notification('Maintenance', market + ' seems to be on an automatic maintenance. Will try every 5 minutes.')
        except: 
            lprint(["Failed to send notification"])    
        while ticker_upd is None: 
            time.sleep(300)  
            lprint(["Market could be on maintenance. Sleeping for 5 minutes."])    
            try:
                ticker_upd = getticker(exchange, market) 
            except: 
                ticker_upd = None
            price_upd = ticker_upd
            price_extreme = price_upd
            
    return price_extreme

##################### Candle analysis; returns high, low, and whether the price crossed a value - among N-min intervals (candles) 
def candle_analysis(cross_target): 
    global market, candle_steps, candle_sleep
    ticker_upd = {}
    price_upd = 0
    price_h = 0
    price_l = 0 
    crossed_flag = False
    failed_attempts = 0
    
    for i in range(1, candle_steps + 1): # 5 min: 100 checks x 3 sec (better indication than 30 checks x 10 sec) 
        try:
            ticker_upd = getticker(exchange, market) 
            price_upd = ticker_upd
            if (price_l == 0) or (price_upd < price_l): 
                price_l = price_upd
            if (price_h == 0) or (price_upd > price_h): 
                price_h = price_upd
            if price_upd >= cross_target: 
                crossed_flag = True 
        except:
            #print "Issues with URL (!) for market", market
            failed_attempts += 1
        time.sleep(candle_sleep) 
        
    # Logging failed attempts number
    if failed_attempts > 0: 
        lprint(["Failed attempts to receive price:", failed_attempts])    
    
    # If retreiving prices fails completely
    if failed_attempts == steps_ticker:     
        ticker_upd = None  
        # Could be related to maintenance
        try:
            send_notification('Maintenance', market + ' seems to be on an automatic maintenance. Will try every 5 minutes.')
        except: 
            lprint(["Failed to send notification"])    
        while ticker_upd is None: 
            time.sleep(300)  
            lprint(["Market could be on maintenance. Sleeping for 5 minutes."])    
            try:
                ticker_upd = getticker(exchange, market) 
            except: 
                ticker_upd = None
            price_upd = ticker_upd
            # If unsuccessful
            price_l = 0
            price_h = 0
    return price_l, price_h, crossed_flag
    
##################### Checking if we need to stop buyback
def check_bb_flag():
    global market 
    global db, cur, bb_id
    
    sell_initiate = False 
    sql_string = "SELECT abort_flag FROM bback WHERE id = {}".format(bb_id)
    rows = query(sql_string)
    try: 
        bb_flag = rows[0][0] # first result 
    except: 
        bb_flag = 0  
    return bool(bb_flag)

##################### Looking for rebuy points (buyback), based on 4H candles price action or simpler price action depending on data availability
def buy_back(price_base): 
    global bb_id, market, exchange_abbr
    global td_data_available, start_time, bars, strategy, time_bb_initiated # bars actually need to be recalculated as 1h is used for buyback
    
    if strategy == 'btc': 
        diff_threshold = 0.005 # threshold as a low of the previous 4H minus 0.5% 
    else: 
        diff_threshold = 0.01 # 1% for alts 
    
    flag_reb_c = True 
    td_first_run = True 
    bback_result = False 

    if td_data_available != True: # using a simple 5-min candles analysis if there is no 4H price data 
        # print "Base: ", price_base    #DEBUG 
        price_l_arr = np.zeros(5)        #5x5-min candlesticks
        price_h_arr = np.zeros(5)       #5x5-min candlesticks
        crossed_arr = np.bool(5)        # for crossed
        lprint([market, ': filling the price array'])
        
        # Filling the prices array with 5x5min candles
        while 0 in price_l_arr: 
            price_l, price_h, crossed_flag = candle_analysis(price_base)
            price_l_arr = np.append(price_l_arr, price_l)
            price_h_arr = np.append(price_l_arr, price_l)
            crossed_arr = np.append(crossed_arr, crossed_flag)
            price_l_arr = np.delete(price_l_arr, [0])
            price_h_arr = np.delete(price_h_arr, [0])
            crossed_arr = np.delete(crossed_arr, [0])
            # print "Lows", price_l_arr, '\nHighs', price_h_arr, '\nCrosses', crossed_arr   #DEBUG

        # Running until need to cancel 
        while flag_reb_c: 
            lows_conf = equal_or_increasing(price_l_arr)                 # Higher or equal lows
            crossed_conf = (True in crossed_arr)                            # Any of candles should cross price_base  
            num_conf = ((price_h_arr >= price_base).sum()) >= 3     # At least 3 of 5 candles highs should be equal or above x
            bback_result = bool(lows_conf * crossed_conf * num_conf)
            lprint([market, ": base", price_base, '| lows holding or higher', lows_conf, '| highs lower than base confirmation:', num_conf, '| crossed flag:', crossed_conf, '| result:', bback_result])

            # Check if we need to cancel 
            stop_bback = check_bb_flag()
            if stop_bback: 
                bback_result = False 
                flag_reb_c = False 
            
            # If we need to exit to proceed with buyback
            if bback_result == True:
                lprint([market, ": initiating buyback"])
                flag_reb_c = False 
                
            # Get new values 
            price_l, price_h, crossed_flag = candle_analysis(price_base)
            price_l_arr = np.append(price_l_arr, price_l)
            price_h_arr = np.append(price_l_arr, price_l)
            crossed_arr = np.append(crossed_arr, crossed_flag)
            price_l_arr = np.delete(price_l_arr, [0])
            price_h_arr = np.delete(price_h_arr, [0])
            crossed_arr = np.delete(crossed_arr, [0])
            
            # Updating DB
            if bb_id is not None: 
                sql_string = "UPDATE bback SET curr_price = {} WHERE id = {}".format(price_h, bb_id) 
                rows = query(sql_string)
            
    # If there is detailed 4H data available 
    else: 
        # Update to set stops according to 4H candles and TD 
        if td_first_run: 
            td_first_run = False 
            bars = td_info.stats(market, exchange_abbr, '4h', 50000, 10)     
            time_hour = time.strftime("%H")
     
        while flag_reb_c: 
            # Checking the need to update 
            time_hour_update = time.strftime("%H")
            if (time_hour_update <> time_hour): 
                # Updating the current hour and the TD values 
                time_hour = time_hour_update
                bars = td_info.stats(market, exchange_abbr, '4h', 50000, 5)     
        
            # Check if we need to cancel 
            stop_bback = check_bb_flag()
            if stop_bback: 
                bback_result = False 
                flag_reb_c = False 
            
            # Checking time elapsed from the start of buyback 
            time_elapsed = (math.ceil(time.time() - time_bb_initiated ))/60    

            # Getting the current price 
            price_upd = get_last_price(market)
            lprint([  "TD setup:", bars['td_setup'].iloc[-1], "TD direction:", bars['td_direction'].iloc[-1], "Time elapsed (min):", time_elapsed, "Current price:", price_upd ])        

            # Updating DB
            if bb_id is not None: 
                sql_string = "UPDATE bback SET curr_price = {} WHERE id = {}".format(price_upd, bb_id) 
                rows = query(sql_string)
            
            # Checking if we should buy back: this happens when the price is above a bullish setup candle 
            if (bars['td_direction'].iloc[-1] == 'up') and (time_elapsed > 60) and (price_upd > (bars['high'].iloc[-1])*(1 + diff_threshold)):  
                bback_result = True 
                flag_reb_c = False 
                lprint(["TD buyback initiated"])
            
    # Finishing up 
    return bback_result   
    
##################### Cancelling active orders on particular market if there are any 
def cancel_orders(market):    
    my_orders = getopenorders(exchange, market)
    # print "Orders", my_orders #DEBUG 
    if my_orders <> '': 
        for val in my_orders:
            lprint(["Cancelling open order:", val['OrderUuid'], "quantity", val['Quantity'], 
                   "quantity remaining", val['QuantityRemaining'], "limit", val['Limit'], "price", val['Price']
                   ])
            cancel_stat = cancel(exchange, market, val['OrderUuid'])
            # Wait for a moment if needed 
            # time.sleep(1)

##################### Update information on performed orders
def sell_orders_info():
    global simulation, main_curr_from_sell, commission_total, alt_sold_total, orders_start, no_sell_orders, market, limit_sell_amount
    
    # Updating order history to collect information on new orders (disable in the simulation mode)
    # Further speed improvement would be to change the structure to a proper dict here right away    

    if simulation != True: 
        # Reset values if we are not simulating
        main_curr_from_sell = 0     
        commission_total = 0        
        alt_sold_total = 0 
        
        # Getting information on sell orders executed
        orders_opening_upd = getorderhistory(exchange, market) 
        for elem in orders_opening_upd: 
            orders_new.add(elem['OrderUuid'])
        orders_executed = orders_new.symmetric_difference(orders_start) 
 
        if orders_executed == set([]):
            lprint(["No sell orders executed"])
            no_sell_orders = True 
        else:
            lprint(["New executed orders"])  
            
            for elem in orders_executed: 
                order_info = getorder(exchange, market, elem)               
                if exchange == 'bitmex':         
                    if order_info['Status'] != 'Canceled': 
                        main_curr_from_sell += order_info['simpleCumQty'] 
                    commission_total += 0 
                else: 
                    main_curr_from_sell += order_info['Price']  
                    commission_total += order_info['CommissionPaid']
                qty_sold = order_info['Quantity'] - order_info['QuantityRemaining'] 
                alt_sold_total += qty_sold
                
                lprint([">", elem, "price", order_info['Price'], "quantity sold", qty_sold ]) #DEBUG 
            lprint(["Total price", main_curr_from_sell, "alts sold total", alt_sold_total]) #DEBUG
    else:
        # If the simulation is True - main_curr_from_sell will have simulated value and the commission would be zero. Updating quantity. 
        alt_sold_total = limit_sell_amount
            
##################### Sell orders outcome 
def sell_orders_outcome():
    global no_sell_orders, total_gained, main_curr_from_sell, value_original, commission_total, total_gained_perc, market
    global price_exit, contracts_start # to use in buyback
    
    if no_sell_orders != True: 
        # Calculating totals 
        total_gained = float(main_curr_from_sell) - float(value_original) - float(commission_total)
        
        # Here division by zero error handling
        if float(value_original)  > 0: 
            total_gained_perc = 100*float(total_gained)/float(value_original)   
        else: 
            total_gained_perc = 0 
        
        if total_gained_perc < 0: 
            txt_result = 'lost'  
        else: 
            txt_result = 'gained'
        
        # Average exit price (value/quantity)
        if exchange == 'bitmex':    
            price_exit = float(contracts_start)/float(main_curr_from_sell)   # for bitmex, calculation is done through contracts  
        else: 
            price_exit = float(main_curr_from_sell)/float(alt_sold_total)
            
        percent_gained = str(round(total_gained_perc, 2))
        trade_time = strftime("%Y-%m-%d %H:%M", localtime())
        
        lprint(['Total from all sales', main_curr_from_sell, 'total commission', commission_total])
        lprint(['Profit ', total_gained, ':', round(total_gained_perc, 2), '%']) 
        # Send the notification about results
        send_notification('Trade finished: ' + str(round(total_gained_perc, 2)) + '% ' + txt_result, market + ': Total ' + str(trade) + ' gained from all sales: ' + 
                          str(main_curr_from_sell) + '. Commission paid: ' + str(commission_total) + '. Trade outcome: ' + percent_gained + '% ' + txt_result + '.')
        # Update the xls register 
        try:
            wb = load_workbook("Trade_history.xlsx")
            ws = wb['BOT']
            new_line = [trade_time, trade, currency, alt_sold_total, price_curr, price_exit, main_curr_from_sell, total_gained, percent_gained, simulation]
            ws.append(new_line)
            max_row = ws.max_row
            # Apply a style 
            index_row = "{}:{}".format(max_row, max_row) 
            for cell in ws[index_row]:
                cell.font = Font(name='Arial', size=10)
            wb.save("Trade_history.xlsx")
            
            #if platform_run != 'Windows':  #uncomment if needed 
            #    copyfile('/home/illi4/Robot/Trade_history.xlsx', '/mnt/hgfs/Shared_folder/Trade_history.xlsx')
            
        except: 
            lprint(['Trade history xls unavailable']) 

##################### Setting stop loss based on price data
def stop_reconfigure(mode = None): 
    global db, cur, job_id
    global time_hour
    global market, exchange_abbr, strategy 
    global price_entry
    
    sl_target_upd = None 
    sl_upd = None 
    sl_p_upd = None  
   
    # Stop level depending on the strategy 
    if strategy == 'btc': 
        down_contingency = 0.005    # 0.5% lower than the previous bullish 4H candle   
    else: 
        down_contingency = 0.01     # 1% lower than the previous bullish 4H candle   
    
    time_hour_update = time.strftime("%H")
    if (time_hour_update <> time_hour) or mode == 'now': 
        # Updating the current hour and the TD values 
        time_hour = time_hour_update
        bars_4h = td_info.stats(market, exchange_abbr, '4h', 50000, 5)     
        if bars_4h['td_direction'].iloc[-1] == 'up': 
            sl_target_upd = bars_4h['low'].iloc[-1] * (1 - down_contingency)   
            sl_upd = round(sl_target_upd/price_entry , 5) 
            sl_p_upd = (1.0 - sl_upd)*100.0 
            lprint(["New stop loss level:", sl_target_upd])
        else:    
            lprint(["No bullish 4H candle to update the stop loss"])  
    return sl_target_upd, sl_upd, sl_p_upd
            
            
##################### Mooning trajectory procedure
##################### currently works in the same way as the main cycle when 4H price data is available       
def to_the_moon(price_reached):     
    # Global variables used 
    global main_curr_from_sell, value_original, price_curr, commission_total, price_target, t_m_id, approved_flag, offset_check, comission_rate
    global sleep_timer
    global db, cur, job_id
    global stopped_price
    global trailing_stop_flag, start_time, bars, strategy, diff_threshold
    global sl, sl_target, sl_p 

    sale_trigger = False # default
    
    # Thresholds for post-profit fallback for BTC or ALTS, when 4H detailed data is not available 
    if market == 'USDT-BTC': 
        post_sl_level = 0.9745     # Fix of -2.55% for BTC
    else: 
        post_sl_level = 0.975    # Fix of -2.5% for all the post-profit cases (alts)  
        
    price_max = price_reached       # this will be changed mooning forward
    price_cutoff = price_reached * post_sl_level   # to sell on original TP if we fall below 
    if td_data_available: 
        trailing_stop = sl_target 
    else: 
        trailing_stop = price_max * post_sl_level    # to sell on new high * stop loss threshold   
    
    lprint(["Mooning from:", price_max])   
    rocket_flag = True
    
    # Running the loop 
    while rocket_flag:  
        # Update to set stops according to 4H candles and TD 
        if td_data_available: 
            sl_target_upd, sl_upd, sl_p_upd = stop_reconfigure()
            if sl_target_upd is not None: 
                trailing_stop = sl_target_upd
                sl = sl_upd
                sl_p = sl_p_upd    
    
        price_last_moon = get_last_price(market)
        increase_info = 100*float(price_last_moon - price_target)/float(price_target) 
        lprint(["Price update:", price_last_moon, "higher than original target on", round(increase_info, 2), "%"])

        # Updating the db 
        sql_string = "UPDATE jobs SET price_curr={}, percent_of={}, mooning={} WHERE job_id={}".format(round(price_last_moon, 8), str(round(increase_info, 2)), 1, job_id)
        rows = query(sql_string)
        
        if price_last_moon > price_max: 
            # Setting higher thresholds if there is no 4H data
            price_max = price_last_moon
            if td_data_available == False: 
                trailing_stop = price_max * post_sl_level        
            lprint(["Last price:", price_max, "| trailing stop", trailing_stop, "| original take profit", price_cutoff])

        #  Checking if this is a time to sell now   
        #  starting only when trailing_stop_flag is active (should not be doing this for BTC runs) 
        # print ">> Price last moon (to compare)", price_last_moon, "maximum price", price_max, "price_cutoff", price_cutoff, "trailing_stop", trailing_stop  # DEBUG # 
        
        if trailing_stop_flag: 
            # Simplified this back to basics as we are using the 4H rule and selling if we are falling behind the bullish candle 
            if (price_last_moon <= price_cutoff) or (price_last_moon <= trailing_stop): 
                lprint(["Run out of fuel @", price_last_moon])
                # Check if we need to sell
                sale_trigger = ensure_sale(price_last_moon)   
                lprint(["Sale trigger (post-profit)", sale_trigger])
            
            # It is a good idea to sell at +45% when something is pumping 
            if increase_info > 45:
                sale_trigger = True 
            
            # Now checking sale trigger and selling if required         
            if sale_trigger == True:  
                lprint(["Triggering trailing stop on", price_last])
                send_notification('Sell: Post-TP', market + ': Triggering trailing stop on the level of ' + str(price_last))
                status = sell_now(price_last_moon)
                # Update the status
                rocket_flag, stat_msg = process_stat(status)
                lprint([stat_msg])            
                # For buyback - using rebuy price
                if price_last_moon > price_cutoff: 
                    stopped_price = trailing_stop 
                else: 
                    stopped_price = price_cutoff       
                            
        # Check if 'sell now' request has been initiated
        sell_init_flag = check_sell_flag()
        if sell_init_flag == True:       
            lprint(["Sale initiated via Telegram @", price_last])
            status = sell_now(price_last_moon)
            sql_string = "UPDATE jobs SET selling = 0 WHERE job_id = {}".format(job_id)     # updating the DB 
            rows = query(sql_string)
            
            # Handling results
            rocket_flag, stat_msg = process_stat(status)
            lprint([stat_msg])
            
            # For buyback - using rebuy price
            if price_last_moon > price_cutoff: 
                stopped_price = trailing_stop 
            else: 
                stopped_price = price_cutoff
                
        # Checking Telegram requests and answering 
        if rocket_flag:
            approved_flag = check_cancel_flag()
            if approved_flag == False: 
                lprint(["Shutdown was requested via Telegram"])   
                sleep_timer = 0
            time.sleep(sleep_timer)

        if approved_flag == False:  # aborting if asked          
            status = 'abort_telegram'
            rocket_flag, stat_msg = process_stat('abort_telegram')

    # Finished the loop - returning the proper code
    return status

##################### Anti-manipulation and anti-flash-crash filter    
def ensure_sale(check_price): 
    proceed_sale = False           # default 
    price_arr = np.zeros(3)         # 3 * N-min candlesticks  (see candle_extreme for N) 
    lprint(["Running ensure_sale check"])
    
    # Filling the prices array - will be checking for lower highs 
    while (0 in price_arr):  
        approved_flag = check_cancel_flag() # checking Telegram requests and answering 
        if approved_flag == False: 
            break
    
        price_highest = candle_extreme('H')  
        price_arr = np.append(price_arr, price_highest)
        price_arr = np.delete(price_arr, [0])
        
        # Selling on the series of lower or same highs of 3 x N-min candlesticks when the price array is filled      
        if (0 not in price_arr): 
            lprint(["High in the candle:", price_highest, "| lower or same highs:", equal_or_decreasing(price_arr)])  #lprint([price_arr]) # DEBUG
            if equal_or_decreasing(price_arr): 
                proceed_sale = True
                break
        else: 
            lprint(["High in the candle:", price_highest])  #lprint([price_arr]) # DEBUG
            
        # If we are back above the check_price value - exit the cycle and return false 
        if price_highest > check_price: 
            lprint(["Cancelling ensure_sale since the price is back to normal"])  
            proceed_sale = False
            break
        
    # Common sense  
    if (price_arr.min() < (price_curr * flash_crash_ind)) and (0 not in price_arr):
        lprint(["Ridiculously low price, better check it."])
        chat.send(market +": ridiculously low price, better check it")
        proceed_sale = False 
        
    return proceed_sale     

##################### Main sell function to sell at current prices   
# Will be performed until the balance available for sale is zero or slightly more      
def sell_now(at_price):
    
    # To decrease price gradually compared to the last average sell price if orders are not filled. Start with zero (percent), maximum 5%
    decrease_price_step = 0.0 
    decrease_attempts_total = 0 
    # First run flag now to sleep on the first call 
    proceed_w_sleep = False
    
    # Global variables used 
    global main_curr_from_sell, value_original, price_curr, commission_total, simulation, currency, market, t_m_id, approved_flag, offset_check, simulation_balance, sell_portion, limit_sell_amount, comission_rate, exchange
    global sleep_sale, steps_ticker, sleep_ticker
    global db, cur, job_id
    global chat
    global balance_start, contracts_start
    
    # Starting balance for further use. Should be done with all orders cancelled
    cancel_orders(market)
    
    # Get balance
    if simulation != True: 
        balance = getbalance(exchange, currency)
        balance_start  = Decimal('{0:.8f}'.format(balance['Available']))   # to correctly work with decimal numbers; not needed for bitmex 

        if exchange != 'bitmex':         
            lprint(["Balance available to sell", balance_start])    #DEBUG
    
    if limit_sell_amount is not None: 
        limit_sell_amount = Decimal(str(limit_sell_amount))     # using str, we will not have more decimal numbers than needed
    if sell_portion is not None: 
        sell_portion = Decimal(str(sell_portion))  
    
    if simulation == True: 
        balance_start = Decimal(str(simulation_balance))
        balance_available = Decimal(str(simulation_balance))
        remaining_sell_balance = Decimal(str(simulation_balance))
        
    # Limiting if required. Should be done with orders cancelled
    if (limit_sell_amount < balance_start) and (limit_sell_amount > 0):
        balance_adjust = Decimal(str(balance_start)) - Decimal(str(limit_sell_amount))
        balance_start = Decimal(str(limit_sell_amount))
        #print ">> Adjust", balance_adjust, "Bal_start", balance_start, "Limit sell am", limit_sell_amount      #DEBUG 
        lprint(["Limiting total amount to be sold. Total:", limit_sell_amount, "Adjustment:", balance_adjust])
    else:
        balance_adjust = 0

    # For bitmex, we will be trading contracts, no adjustments are available. Getting the balances and setting the original value 
    if exchange == 'bitmex': 
        contracts_check = getpositions(exchange, market)[0]
        if contracts_check == {}: 
            sell_run_flag = False
            contracts = 0
        else: 
            contracts = contracts_check['contracts'] 
            contracts_start = contracts
            balance_available = contracts
            balance_adjust = 0 
            sell_portion = balance_available
        # Original value 
        value_original = Decimal(str(contracts_check['xbt']))
    else: # for other exchanges     
        value_original = Decimal(str(price_entry)) * balance_start    
 
    lprint(["Original value:", value_original])
    
    # Main sell loop
    sell_run_flag = True
    stopmessage = 'stop' # default stop message meaning successful sale
    
    while sell_run_flag: 
        decrease_price_flag = False     # Reset the overall flag to decrease price 
      
        # Wait until existing orders are cancelled - that is why we need sleep here and not in the end 
        # Checking Telegram requests and cancelling if needed
        if proceed_w_sleep: 
            time.sleep(sleep_sale)
        
        # 0. Check open orders, cancel if unfilled, and decrease price further compared to average last 
        my_orders = getopenorders(exchange, market)
        if my_orders <> '': 
            for val in my_orders:
                # Checking if some are open not filling
                if (val['Quantity'] == 0):
                    unfilled_prop = 0
                else:
                    unfilled_prop = Decimal(str(val['QuantityRemaining']))/Decimal(str(val['Quantity']))
                if unfilled_prop >= 0.05:  # if more than 5% still left in the order
                    lprint(["Cancelling unfilled order:", val['OrderUuid'], "quantity", val['Quantity'], 
                           "quantity remaining", val['QuantityRemaining'], "limit", val['Limit'], "price", val['Price']
                           ]) 
                    cancel_stat = cancel(exchange, market, val['OrderUuid'])
                    time.sleep(5) # Wait for cancellations to be processed just in case 
                    # Then we will get information on available balance which includes cancellations
                    # Set decrease price flag
                    decrease_price_flag = True
                    decrease_attempts_total += 1

        # Decrease price more compared to last prices if required
        if (decrease_price_step < 0.05) and decrease_price_flag:
            decrease_price_step += 0.005
            lprint(["Sell price will be decreased on", decrease_price_step*100, "%"]) 
            
        # Notify if a position cannot be sold for a long time 
        if decrease_attempts_total >= 30: 
            time_passed = int(decrease_attempts_total*(sleep_sale + steps_ticker*sleep_ticker)/60)
            lprint(["Unable to sell the position for more than", time_passed, "minutes"]) 
            chat.send(market +": unable to sell the position for more than " + time_passed + " minutes")
                        
        # 1. Get the available balance and proceed with selling       
        if simulation != True: 
            balance = getbalance(exchange, currency)
            balance_available = Decimal('{0:.8f}'.format(balance['Available']))
            # print ">> Balance_available", balance_available #DEBUG 
        else:
            # If we are in the simulation mode - use the value from the previous run
            balance_available = remaining_sell_balance           
        
        # For bitmex, we will be trading contracts, no adjustments are available 
        if exchange == 'bitmex': 
            contracts_check = getpositions(exchange, market)[0]
            if contracts_check == {}: 
                sell_run_flag = False
            else: 
                contracts = contracts_check['contracts'] 
                balance_available = contracts
                balance_adjust = 0 
                sell_portion = balance_available
        else: # for the other exchanges 
            #print ">> Balance_available pre", balance_available    #DEBUG  
            #print  ">> Balance_adjust pre", balance_adjust     #DEBUG  
            
            # Adjusting according to the limit 
            balance_available = balance_available - Decimal(str(balance_adjust))
            if sell_portion == None: 
                sell_portion = balance_available           

        # Check if we have sold everything 
        if balance_available <= balance_start * Decimal(0.01):
            sell_run_flag = False
        
        # Error strings for exchanges 
        err_1 = 'DUST_TRADE_DISALLOWED_MIN_VALUE_50K_SAT'
        err_2 = 'MIN_TRADE_REQUIREMENT_NOT_MET'
        
        # 2. If something is still required to be sold
        if sell_run_flag: 
            lprint(["Sell amount", balance_available, "at price threshold", at_price, "split on", sell_portion])
            remaining_sell_balance = balance_available   
            if exchange == 'bitmex': 
                sale_steps_no = 1
            else: 
                sale_steps_no = int(math.ceil(round(Decimal(str(balance_available))/Decimal(str(sell_portion)), 3)))   
            #print ">> Sell amount", balance_available, "remaining_sell_balance", remaining_sell_balance  #DEBUG#
            
            # Selling loop 
            for i in range(1, sale_steps_no + 1):                
                # Check how much should we sell at this step
                if sell_portion > remaining_sell_balance: 
                    sell_q_step = remaining_sell_balance
                else:
                    sell_q_step = sell_portion
                
                # Price update
                price_last_sell = get_last_price(market)
                # Decreasing the price if necessary
                price_to_sell = price_last_sell*(1 - decrease_price_step)
                lprint(["Placing SELL order: Q:", sell_q_step, "@", price_to_sell, "Last market price:", price_last_sell, 
                       "Remaining balance after sale:", round(remaining_sell_balance - sell_q_step, 6)])
                
                # Actually place sell orders if we are not in the simulation mode - re-check
                if simulation != True: 
                    # For bitmex, we will be placing contracts in the other direction (short)
                    if exchange == 'bitmex': 
                        # balance_available is the number of contracts here 
                        price_to_sell = round(price_to_sell, 0)
                        sell_result = selllimit(exchange, market, sell_q_step, price_to_sell, balance_available) 
                        # print "selllimit({}, {}, {}, {}, {})".format(exchange, market, sell_q_step, price_to_sell, balance_available) #DEBUG 
                    else: 
                        sell_result = selllimit(exchange, market, sell_q_step, price_to_sell) 
                    
                    lprint([">> Sell result:", sell_result])  # DEBUG # 
                    
                    if (sell_result == err_1) or (sell_result == err_2):
                        sell_run_flag = False
                        stopmessage = 'err_low'
                    else:
                        # Checking if the sell order was placed
                        try: 
                            if 'uuid' not in sell_result.keys():
                                # Issue with placing order
                                # DEBUG # print "Issue"
                                sell_run_flag = False
                                stopmessage = 'no_idea'
                        except:
                            # DEBUG # print "Issue"
                            sell_run_flag = False
                            stopmessage = 'no_idea'
                
                else: 
                    # If in simulation - calculate profit from virtual sale.  
                    main_curr_from_sell += float(sell_q_step) * price_to_sell 
                    commission_total += float(sell_q_step)*price_to_sell * comission_rate

                # Update the db with price_last_sell
                sql_string = "UPDATE jobs SET price_curr={}, selling={} WHERE job_id={}".format(round(price_last_sell, 8), 1, job_id)
                rows = query(sql_string)

                # Decrease remaining balance to sell 
                remaining_sell_balance = remaining_sell_balance - sell_q_step

        # Checking Telegram requests and answering 
        approved_flag = check_cancel_flag()
        if approved_flag == False: 
            # Aborting if asked
            sell_run_flag = False
            stopmessage = 'abort_telegram'
        # Change the flag to sleep on the next cycle
        proceed_w_sleep = True    
        
    # Finishing up
    #print "main_curr_from_sell {}, commission_total {}, contracts_start {}".format (main_curr_from_sell,  commission_total, contracts_start) # DEBUG 
    
    return stopmessage

################################ Functions - system ############################################
def terminate_w_message(short_text, errtext):
    global logger, handler    
    lprint([short_text])
    send_notification(short_text, errtext)
    logger.close_and_exit()

# Checking if we need to terminate
def check_cancel_flag():
    global market 
    global db, cur, job_id
    
    keep_running = True 
    sql_string = "SELECT abort_flag FROM jobs WHERE job_id = '{}'".format(job_id)
    rows = query(sql_string)

    try: 
        flag_terminate = rows[0][0] # first result 
    except: 
        flag_terminate = 0
    if (flag_terminate == 1): 
        keep_running = False
    return keep_running
 
# Checking if we need to initiate selling from the main or from the mooning cycle 
def check_sell_flag():
    global market 
    global db, cur, job_id
    
    sell_initiate = False 
    sql_string = "SELECT selling FROM jobs WHERE market = '{}'".format(market)
    rows = query(sql_string)

    try: 
        sell_flag = rows[0][0] # first result 
    except: 
        sell_flag = 0
    if (sell_flag == 1): 
        sell_initiate = True
    return sell_initiate
 
def send_notification(subj, text):
    global send_messages, trade_id, comm_method, market
    
    if send_messages:
        if comm_method == 'mail':
            msg = MIMEMultipart()
            msg['From'] = fromaddr
            msg['To'] = toaddr
            msg['Subject'] = trade_id + ': ' + subj
            body = text
            msg.attach(MIMEText(body, 'plain'))
            server = smtplib.SMTP('smtp.gmail.com', 587)
            server.starttls()
            server.login(fromaddr, email_passw)
            text = msg.as_string()
            server.sendmail(fromaddr, toaddr, text)
            server.quit()  
        else: 
            chat.send(text)
            
def timenow():
    return strftime("%Y-%m-%d %H:%M:%S", localtime())

###################################################################################
############################## Main workflow #########################################
###################################################################################

if stop_loss: 
    lprint([market, "| Take profit target:", price_target, "| Stop loss:", sl_target, "| Simulation mode:", simulation])
else: 
    lprint([market, "| Take profit target:", price_target, "| Stop loss: disabled (post-profit only) | Simulation mode:", simulation])

if limit_sell_amount > 0: 
    lprint(["Maximum quantity to sell", limit_sell_amount])

'''
######## Removed this - will not really be using SL (at least for now)
# Check if TP is set higher than SL 
if tp < sl: 
    # print "TP {}, SL {}".format(tp, sl) # DEBUG #
    lprint(["Take profit lower than stop loss, r u ok?"])
    exit(0)
''' 

time_hour = time.strftime("%H")     # For periodic updates of 4H candles and stops 
  
# 1. Checking market correctness and URL validity, as well as protecting from fat fingers
try: 
    ticker_upd = getticker(exchange, market) 
    # Ticker could be failing if there is automatic maintenance - then sleep for a while
    if ticker_upd is None: 
        send_notification('Maintenance', market + ' seems to be on an automatic maintenance. Will try every 5 minutes.')
        while ticker_upd is None: 
            lprint(["Market could be on maintenance. Sleeping for 5 minutes."])    
            time.sleep(300) # sleeping for 5 minutes and checking again
            ticker_upd = getticker(exchange, market) 
        
    if ticker_upd == 'INVALID_MARKET': 
        lprint(['Error: Invalid market'])
        logger.close_and_exit()

    else:
        # Fat fingers protection    
        price_check = ticker_upd
        ratio = float(price_target)/float(price_check)
        if (ratio >= 8) or (ratio <= 0.15): 
            lprint(['Error: Double-check prices, are you missing a zero or adding an extra one? The current price is', price_check])
            logger.close_and_exit()

except urllib2.URLError:
    terminate_w_message('Exchange url unavailable')
    logger.close_and_exit()
   
### 2. Checking available balance
if simulation != True: 
    balance = getbalance(exchange, currency)
    if balance['Available'] == 0: 
        terminate_w_message('Error: Zero balance', currency + ': zero balance')
        logger.close_and_exit()

### 3. Start the main workflow
run_flag = True 
approved_flag = True
no_sell_orders = False      # default value to suit both simulation and the real run

### 4. Inserting in the sqlite db if started fine ##
sql_string = "INSERT INTO jobs(market, tp, sl, simulation, mooning, selling, price_curr, percent_of, abort_flag, stop_loss, entry_price, mode, tp_p, sl_p, exchange) VALUES ('{}', {}, {}, {}, {},  {},  {},  {},  {}, {}, {}, '{}', {}, {}, '{}')".format(
    market.upper(), price_target, sl_target, int(simulation), int(False), int(False), price_curr, 100, int(False), int(stop_loss), price_entry, simulation_param, tp_p, sl_p, exchange)    
job_id, rows = query_lastrow_id(sql_string)

### 5. Price data for time analysis and strategy. Check what's up with TD analysis data 
start_time = time.time()
td_data_available = True  # default which will be changed to False when needed  
try: 
    bars = td_info.stats(market, exchange_abbr, '4h', 10000, 10)    
    try: 
        if bars == None: 
            td_data_available = False 
    except: 
        for elem in bars['td_setup'][-3:]:      # should have at least 3 bars with filled TD values
            if elem is None: 
                td_data_available = False 
        num_null = bars['open'].isnull().sum()
        if num_null > 0: 
            td_data_available = False 
except: 
    td_data_available = False 
    
print "TD data availability:", td_data_available

### 6. Strategy and thresholds (for non-4H based action) 
if currency in ['XMR', 'DASH', 'ETH', 'LTC', 'XMR']: 
    strategy = 'alt-med'
    diff_threshold = 0.045
elif currency == 'BTC': 
    strategy = 'btc'
    diff_threshold = 0.0255
else: 
    strategy = 'alt-volatile' 
    diff_threshold = 0.055

### 7. 4H-based stop loss update    
if td_data_available: 
    lprint(["Reconfiguring stop loss level based on 4H candles"])
    sl_target_upd, sl_upd, sl_p_upd = stop_reconfigure('now')
    if sl_target_upd is not None: 
        sl_target = sl_target_upd
        sl = sl_upd
        sl_p = sl_p_upd    

### 8. Creating new set to store previously executed orders. Will be used to calculate the gains 
orders_start = set()
orders_new = set()

orders_opening = None   #sometimes api fails and ends with an error - so retrying here
while orders_opening is None:
    try:
        orders_opening = getorderhistory(exchange, market)
    except:
         time.sleep(1) 
 
lprint(["Last orders when starting the script"])
if len(orders_opening) < 5: 
    count_max = len(orders_opening)
else: 
    count_max = 5 

for i in range (0, count_max): 
    lprint(['>', orders_opening[i]['OrderUuid']])  
 
for elem in orders_opening: 
    orders_start.add(elem['OrderUuid'])
    #lprint(['>', elem['OrderUuid']]) #DEBUG

# Flags to notify if the prices dropped 
flag_notify_m = True
flag_notify_h = True
dropped_flag = False
    
### 9. Start the main cycle
while run_flag and approved_flag:  
    try:    # try / except is here to raise keyboard cancellation exceptions
        if td_data_available:         # update the stop loss level if due and if we have data
            sl_target_upd, sl_upd, sl_p_upd = stop_reconfigure()
            if sl_target_upd is not None: 
                sl_target = sl_target_upd
                sl = sl_upd
                sl_p = sl_p_upd    
                sql_string = "UPDATE jobs SET sl={}, sl_p={} WHERE job_id={}".format(sl_target, sl_p, job_id)   # updating the DB 
                rows = query(sql_string)
            
        # Get the last price
        price_last = get_last_price(market)
        price_compared = round((float(price_last)/float(price_curr))*100, 2)
        lprint([market, ": updating price information:", price_last, "|", price_compared, "% of entry price | sl:", sl_target ])
        sql_string = "UPDATE jobs SET price_curr={}, percent_of={} WHERE job_id={}".format(round(price_last, 8), price_compared, job_id)
        rows = query(sql_string)
        
        # Running the main conditions check to trigger take profit / stop loss 
        if price_last >= price_target:  # price target reached, notify and start mooning 
            lprint(["Take-profit price reached"])
            send_notification("Mooning", market + ": Reached the initial TP target and mooning now: " + str(price_last))
            status = to_the_moon(price_last)    # mooning for as long as possible 
            if status == 'stop':
                lprint(["Stopped monitoring and finished trades (post-profit)"])
                sleep_timer = 0
                run_flag = False 
                stopped_mode = 'post-profit'    # used in buyback
            elif status == 'abort_telegram': 
                lprint(["Stopped monitoring and finished trades (as requested)"])
                sleep_timer = 0
                run_flag = False 
                stopped_mode = 'telegram' 

        # Broken lower level (stop loss) if stop loss is enabled 
        if stop_loss: 
            if price_last <= sl_target: 
                dropped_flag = True     # changing the flag 
                lprint(["Hitting pre-profit stop loss threshold:", sl_target])
                sale_trigger = ensure_sale(sl_target)   # check if we need to sell 
                lprint(["Sale trigger (pre-profit):", sale_trigger])
                if sale_trigger == True:       
                    # Stop-loss triggered
                    lprint(["Triggering pre-profit stop loss on", price_last])
                    send_notification('Sell: SL', market + ': Triggering pre-profit stop loss at the level of ' + str(price_last))
                    status = sell_now(price_last)
                    # Handling results
                    run_flag, stat_msg = process_stat(status)
                    lprint([stat_msg])
                    stopped_mode = 'pre-profit'     # used in buyback
                    stopped_price = sl_target
 
        # Check if selling now request has been initiated
        sell_init_flag = check_sell_flag()
        if sell_init_flag and approved_flag and run_flag:       
            lprint(["Sale initiated via Telegram @", price_last])
            status = sell_now(price_last)
            # Handling results
            run_flag, stat_msg = process_stat(status)
            lprint([stat_msg])
            stopped_price = price_last  # used in buyback
        
        # Checking cancellation request and sleeping 
        if run_flag and approved_flag:
            approved_flag = check_cancel_flag()
            if approved_flag == False: 
                lprint(["Shutdown was requested via Telegram"])   
                sleep_timer = 0
            time.sleep(sleep_timer)
            
    except KeyboardInterrupt:
        lprint(["Shutdown was initiated manually, canceling orders and terminating now"])   
        sql_string = "DELETE FROM jobs WHERE job_id = {}".format(job_id)    # deleting the task from the db 
        rows = query(sql_string)

        # Cancelling orders if not in the simulation mode
        if simulation != True:
            cancel_orders(market)
            time.sleep(10) # wait for cancellations to be processed
            # Information on orders performed if not in a simulation mode
            sell_orders_info()
            sell_orders_outcome()
        logger.close_and_exit()


### 10. Exit point for the main cycle, sell cycle, mooning cycle 
sql_string = "DELETE FROM jobs WHERE job_id = {}".format(job_id)    # deleting the task from the db 
rows = query(sql_string)

# Just a simulation and cancelled by Telegram thing - no virtual sell orders
if simulation == True and approved_flag != True:   
    no_sell_orders = True

### 11. Getting information on performed sell orders and displaying / recording the outcomes
sell_orders_info()
sell_orders_outcome()

# Then start monitoring for buyback (both post-moon and SL)  'pre-profit'  /  'post-profit'
''' 
# Uncomment if you would like to restrict buybacks 
# Checking losses to stop buyback in case of 2 consecutive losses incurred  
sql_string = "SELECT id FROM losses WHERE market = '{}'".format(market)
rows = query(sql_string)
try: 
    loss_id = int(rows[0][0]) # if one loss already have been incurred - no bback 
except: 
    loss_id = None
''' 
if (stopped_mode == 'pre-profit') and (cancel_buyback == False): 
    bb_price = price_exit * 0.9975  # Using value of price_exit (actual sell price) minus 0.25% (for non-4H action) 
    if td_data_available:  
        lprint(["Buyback will be based on 4H candles"])
    else: 
        lprint(["Setting buyback price as actual sell -0.25%. Price_exit:", price_exit, "bb_price", bb_price])
    
    '''
    # Uncomment if you would like to restrict buybacks 
    # Checking losses to stop buyback in case of 2 consecutive losses incurred  
    if loss_id is not None: 
        chat.send(market +": stopped buyback after two consecutive losses")
        sql_string = "DELETE FROM losses WHERE id = {}".format(loss_id)
        rows = query(sql_string)
        logger.close_and_exit()
        exit(0) 
    else: 
        # Inserting into losses table if this is the first occurence 
        sql_string = "INSERT INTO losses(market) VALUES ('{}')".format(market)
        rows = query(sql_string)    
    ''' 
elif stopped_mode == 'post-profit': 
    # Thresholds for post-profit fallback for BTC or ALTS
    if market == 'USDT-BTC': 
        bb_price = price_exit  * 1.005 # Using fixed value of +0.5% from stop; however, does not refer to price value when using TD analysis 
    else: 
        bb_price = price_exit  * 1.01 
    
    if td_data_available:  
        lprint(["Buyback will be based on 4H candles"])
    else: 
        lprint(["Setting buyback price as actual +1%. Stopped_price:", stopped_price, "bb_price", bb_price])
    ''' 
    # Uncomment if you would like to restrict buybacks 
    # If it was a loser - delete the info in DB and continue with BBack 
    if loss_id is not None: 
        sql_string = "DELETE FROM losses WHERE id = {}".format(loss_id)
        rows = query(sql_string)
    ''' 
else: 
    # If just called for stop from Telegram 
    lprint(["Sold through telegram"])   
    bb_price = price_exit

### 12. Buying back based on 4H action or alternative price action.   
try: 
    lprint(["Buyback monitoring started:", stopped_mode, "| TD data availability:", td_data_available])   
    buy_trade_price = float(balance_start) * bb_price * (1 - comission_rate) # commission depending on the exchange. If we do not have TD data
    
    # Inserting into buyback information table 
    sql_string = "INSERT INTO bback(market, bb_price, curr_price, trade_price, exchange) VALUES ('{}', {}, {}, {}, '{}')".format(market, bb_price, bb_price, buy_trade_price, exchange)
    bb_id, rows = query_lastrow_id(sql_string)      
    
    time_bb_initiated = time.time()     # getting a snapshot of time for buyback so that we wait for at least an hour before starting buyback 
    bb_flag = buy_back(bb_price)      # runs until a result is returned     
    
    # If we have reached the target to initiate a buyback and there was no cancellation through Telegram
    if bb_flag: 
        send_notification('Buyback', 'Buy back initiated for ' + market)  
        
        # Launching workflow to buy and resume the task with same parameters
        # Insert a record in the db: workflow(wf_id INTEGER PRIMARY KEY, tp FLOAT, sl FLOAT, sell_portion FLOAT)
        tp_price = bb_price * tp
        sl_price = bb_price * (1 - diff_threshold)  # depending on the strategy 
        sql_string = "INSERT INTO workflow(tp, sl, sell_portion, run_mode, price_entry, exchange) VALUES ({}, {}, {}, '{}', {}, '{}')".format(tp_price, sl_price, 0, simulation_param, float(bb_price), exchange_abbr)
        wf_id, rows = query_lastrow_id(sql_string)       

        if wf_id is not None: 
            buy_market = '{0}-{1}'.format(trade, currency)
            sql_string = "UPDATE workflow SET market = '{}', trade = '{}', currency = '{}', exchange = '{}' WHERE wf_id = {}".format(market, trade, currency, exchange, wf_id) 
            job_id, rows = query_lastrow_id(sql_string)
            
        # Buy depending on the platform. We will buy @ market price now, and the price entry price is already in the DB
        if td_data_available: 
            mode_buy = 'now' 
        else: 
            mode_buy = 'reg' 

        logger.close() # closing logs 
        
        sql_string = "DELETE FROM bback WHERE id = {}".format(bb_id)  # deleting buyback from the table 
        rows = query(sql_string)
        
        # Run a smart buy task now when we have a buyback confirmation 
        python_call = 'python smart_buy.py ' + ' '.join([mode_buy, exchange_abbr, trade, currency, str(buy_trade_price)])
        p = subprocess.Popen(python_call, shell=True, stderr=subprocess.PIPE)
        while True:
            out = p.stderr.read(1)
            if out == '' and p.poll() != None:
                break
            if out != '':
                sys.stdout.write(out)
                sys.stdout.flush()
    
    # If a buyback cancellation was requested 
    else: 
        send_notification('Buyback', 'Buy back cancelled as requested for ' + market)  
    
except KeyboardInterrupt:
    print "Buyback cancelled or the task was finished"  
    sql_string = "DELETE FROM bback WHERE id = {}".format(bb_id)
    rows = query(sql_string)
    logger.close_and_exit()

