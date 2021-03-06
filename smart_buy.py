##################################################################################################################
# Use: python smart_buy.py mode exchange basic_curr altcoin total_in_basic_curr [price] [time limit for the price in minutes] 
# Example: reg/brk/now/reg-s/brk-s/4h btrx BTC QTUM 0.005 0.0038 15 
# This tries to buy QTUM for 0.005 BTC at Bittrex for the price of 0.0038 for 15 minutes, then switches to market prices
#
# Modes:
#   4h - buy based on 4h candles price action 
#   reg - buy at fixed price 
#   brk - buy on breakout (above the specified price) 
# hint: options above with -s mean the same but they run in the simulation mode
#   now is to buy immediately 
#
# Exchanges: btrx, bina, bmex (bittrex, binance, bitmex) 
#
# For bitmex, negative values can be used. E.g. python smart_buy.py now bmex usd btc -0.3

################################ Libraries ############################################
# Standard libraries 
import time
import sys
from sys import exit, argv
from time import localtime, strftime
import subprocess   
import math
import smtplib
from email.MIMEMultipart import MIMEMultipart
from email.MIMEText import MIMEText
import urllib2
import decimal
from decimal import Decimal
from openpyxl import Workbook, load_workbook   
from openpyxl.styles import Font, Fill
from shutil import copyfile
import os
import numpy as np
import traceback

## Custom libraries 
from telegramlib import telegram                              # my lib to work with Telegram
from sqltools import query_lastrow_id, query           # proper requests to sqlite db
from loglib import logfile # logging
import platformlib as platform                                    # detecting the OS and assigning proper folders 

# Universal functions for all exchanges 
from exchange_func import getticker, getopenorders, cancel, getorderhistory, getorder, getbalance, selllimit, getorderbook, buylimit, getbalances, binance_price_precise, binance_quantity_precise, getpositions, closepositions, bitmex_leverage

# Using coinigy to get prices so that there are no stringent restrictions on api request rates (frequency)
from coinigylib import coinigy 
coinigy = coinigy()

################################ Config - part I ############################################

### Import a configuration file 
import config 

### TD analysis library
import tdlib as tdlib
td_info = tdlib.tdlib()

### Decimal precision and roubding 
decimal.getcontext().prec = 25
decimal.getcontext().rounding = 'ROUND_DOWN'

print "Running..."

### Interval and number of checks to get current (last) prices 
steps_ticker = config.steps_ticker 
sleep_ticker = config.sleep_ticker      # So that ticker in total takes 30 seconds 

### Sleep timer in seconds for buy task. changed to 3 minutes because orders were filling way too quickly at a higher price
sleep_timer = config.buy_sleep_timer   

### Orders sequence for a market price calculation 
orders_check = config.orders_check 

### Steps and timer for buybacks 
candle_steps = config.candle_steps
candle_sleep = config.candle_sleep     # Tested, 3 sec lead to having ~5 min 30 sec in between  

### Set up the speedrun multiplier if need to test with higher speeds. 1 is normal, 2 is 2x faster 
speedrun = config.speedrun

sleep_timer = int(sleep_timer/speedrun)
sleep_ticker = int(sleep_ticker/speedrun)
candle_steps = int(candle_steps/speedrun)

### Comms 
send_messages = True 
comm_method = config.comm_method 
chat = telegram()

### Default flag for shorting. The bot can be used to short on bitmex, not only go long 
short_flag = False 

bitmex_margin = config.bitmex_margin   # size of margin on bitmex, minor for now

# Time analysis candles length 
td_period = config.td_period   # possible options are in line with ohlc (e.g. 1h, 4h, 1d, 3d); customisable. This sets up smaller time interval for dynamic stop losses and buy backs     
td_period_extended = config.td_period_extended   # possible options are in line with ohlc (e.g. 1h, 4h, 1d, 3d); customisable. This sets up larger time interval for buy backs (should be in line with the smaller one)    

### Platform
platform = platform.platformlib()
platform_run, cmd_init, cmd_init_buy = platform.initialise() 

################################ Functions - part I ############################################

##################### Price comparison 
def strictly_increasing(L):
    return all(x<y for x, y in zip(L, L[1:]))

def strictly_decreasing(L):
    return all(x>y for x, y in zip(L, L[1:]))
   
def terminate_w_message(short_text, errtext):
    global logger, handler    
    lprint([short_text])
    send_notification(short_text, errtext)
    logger.close_and_exit()

def send_notification(subj, text):
    global send_messages, comm_method, market, chat
    
    if send_messages:
        if comm_method == 'mail':
            msg = MIMEMultipart()
            msg['From'] = fromaddr
            msg['To'] = toaddr
            msg['Subject'] = trade_id + ': ' + subj
            body = text
            msg.attach(MIMEText(body, 'plain'))
            server = smtplib.SMTP('smtp.gmail.com', 587)
            server.starttls()
            server.login(fromaddr, email_passw)
            text = msg.as_string()
            server.sendmail(fromaddr, toaddr, text)
            server.quit()  
        else: 
            chat.send(text)

##################### Log and print function 
def lprint(arr):
    msg = ' '.join(map(lambda x: ''+ str(x), arr))
    try: 
        logger.write(msg)
        print msg
    except: 
        print 'Failed to print output due to the IO error'
            
################################ Config - part II ############################################

### Input parameters 
try: 
    # Modes 
    mode = argv[1].lower()
    if mode not in ['reg', 'brk', 'now', 'reg-s', 'brk-s', '4h', '4h-s', 'fullta', 'fullta-s']: 
        print "Incorrect mode specified (should be reg, 'reg', 'brk', 'now', 'reg-s', 'brk-s', '4h', '4h-s', 'fullta', 'fullta-s')\n\n"
        send_notification('Incorrect mode', 'Incorrect mode specified')
        exit(0)
        
    # Exchange 
    exchange_abbr = argv[2].lower()
    if exchange_abbr not in ['btrx', 'bina', 'bmex']: 
        print 'Incorrect exchange specified (should be btrx, bina, or bitmex)\n\n'
        send_notification('Incorrect exchange', 'Incorrect exchange specified')
        exit(0)
        
    if exchange_abbr == 'btrx': 
        exchange = 'bittrex' 
        comission_rate = config.comission_rate_bittrex
    elif exchange_abbr == 'bina': 
        exchange = 'binance' 
        comission_rate = config.comission_rate_binance
    elif exchange_abbr == 'bmex': 
        exchange = 'bitmex' 
        comission_rate = config.comission_rate_bitmex

    # Main currency (e.g. BTC) 
    market = argv[3].upper()
    try:
        trade, currency = market.split('-')
    except: 
        trade = market  # e.g. if only one market vs BTC is provided - such as XRPH18 on bitmex  
        currency = 'BTC'
    
    # New logger
    logger = logfile(market, 'buy')    

    # Getting for the whole if there is no input 
    balance = getbalance(exchange, trade)
    balance_trade = float(balance['Available'])
    #print "TRADE BALANCE", balance_trade   #DEBUG

    try: 
        source_position = float(argv[4])
        # For bitmex shorts 
        if source_position < 0: 
            short_flag = True 
            source_position = abs(source_position)   
    except: 
        balance = getbalance(exchange, trade)
        source_position = float(balance['Available'])
        lprint(["Buying for the whole balance of", source_position])
    
    # Also change to properly reflect the margin    
    if exchange == 'bitmex': 
        source_position = source_position*bitmex_margin
    
    # If the price is set up
    try: 
        fixed_price = float(argv[5])
        fixed_price_flag = True
    except:
        fixed_price = 0
        fixed_price_flag = False
        
    # Time restriction for fixed price orders in minutes
    try: 
        time_restriction = float(argv[6])
    except:
        time_restriction = 0

    # Shorts are only supported for bitmex 
    if exchange != 'bitmex' and short_flag: 
        lprint(["Shorts are not supported on the exchange", exchange])
    
    ### Greetings (for logs readability) 
    lprint(["###################### SMART_BUY ###########################"])
        
    
except:
    print 'Specify the parameters: mode exchange basic_curr-altcoin total_in_basic_curr [price] [time limit for the price in minutes] \n>Example: reg/brk/now/reg-s/brk-s/4h btrx BTC-QTUM 0.005 0.0038 15 \nThis tries to buy QTUM for 0.005 BTC at Bittrex for the price of 0.0038 for 15 minutes, then switches to market prices \n\nModes: \n4h - buy based on 4h candles price action \nreg - buy at fixed price \nbrk - buy on breakout (above the specified price) \noptions with -s mean the same but they run in the simulation mode \nnow is immediately \n\nExchanges: btrx, bina, bmex (bittrex, binance, bitmex)'
    exit(0)
    
### Thresholds for buys on 4H 
if currency == 'BTC': 
    diff_threshold = 0.005  # 0.5% 
else: 
    diff_threshold = 0.01   # 1% 

### Set up the margin on bitmex 
if exchange == 'bitmex': 
    set_margin = bitmex_leverage(market, bitmex_margin)
    
### Price data analysis 
time_hour = time.strftime("%H") 
td_data_available = True  # default which will be changed to False when needed  
td_data_extended_available = True  # default which will be changed to False when needed  

# Sleeping for a bit so that information on workflows is updated in the database just in case 
time.sleep(int(30/speedrun))

# TD data availability 
try: 
    bars = td_info.stats(market, exchange_abbr, td_period, 50000, 10)    
    bars_extended = td_info.stats(market, exchange_abbr, td_period_extended, 50000, 10)    
    try: 
        if bars == None: 
            td_data_available = False 
        if bars_extended == None: 
            td_data_extended_available = False 
    except: 
        # Smaller interval 
        for elem in bars['td_setup'][-3:]:      # should have at least 3 bars with filled TD values
            if elem is None: 
                td_data_available = False 
        num_null = bars['open'].isnull().sum()
        if num_null > 0: 
            td_data_available = False 
        # Larger interval 
        for elem in bars_extended['td_setup'][-3:]:      # should have at least 3 bars with filled TD values
            if elem is None: 
                td_data_extended_available = False 
        num_null = bars_extended['open'].isnull().sum()
        if num_null > 0: 
            td_data_extended_available = False 
except: 
    td_data_available = False 
    td_data_extended_available = False 
 
print "TD data availability:", td_data_available, "| extended ", td_data_extended_available

### Check if this is a part of workflow (meaning that a job should be then launched)   
sql_string = "SELECT wf_id, run_mode FROM workflow WHERE market = '{}' AND exchange = '{}'".format(market, exchange_abbr)
rows = query(sql_string)

try: 
    wf_id = rows[0][0]   # first result if existing 
    wf_run_mode = rows[0][1] 
    lprint(["Workflow:", wf_id, wf_run_mode])
except:
    lprint(["Not a part of workflow"])
    wf_id = None
    wf_run_mode = None

################################ Functions - part II ############################################
##################### Check if cancellation was requested through Telegram 
def check_cancel_flag():
    global db, cur, job_id
    keep_running = True 
    sql_string = "SELECT abort_flag FROM buys WHERE job_id = '{}'".format(job_id)
    rows = query(sql_string)
    try: 
        flag_terminate = rows[0][0] # first result 
    except: 
        flag_terminate = 0
    if (flag_terminate == 1): 
        keep_running = False
    return keep_running

##################### Getting several last prices in short intervals instead of just one 
def get_last_price(market): 
    ticker_upd = {}
    price_upd = 0
    failed_attempts = 0
    for i in range(1, steps_ticker + 1):
        try:
            #ticker_upd = getticker(exchange, market) 
            ticker_upd = coinigy.price(exchange_abbr, market)
            price_upd += ticker_upd
        except:
            #print "Issues with URL (!) for market", market
            failed_attempts += 1
        time.sleep(sleep_ticker)
        
    # Logging failed attempts number
    if failed_attempts > 0: 
        lprint(["Failed attempts to receive price:", failed_attempts])   
        
    # If retreiving prices fails completely
    if failed_attempts == steps_ticker:     
        ticker_upd = None # Change
        # Could be related to maintenance
        try:
            send_notification('Maintenance', market + ' seems to be on an automatic maintenance. Will try every 5 minutes.')
        except: 
            lprint(["Failed to send notification"])    
        while ticker_upd is None: 
            time.sleep(300) # sleeping for 5 minutes and checking again
            lprint(["Market could be on maintenance. Sleeping for 5 minutes."])    
            try:
                #ticker_upd = getticker(exchange, market)  
                ticker_upd = coinigy.price(exchange_abbr, market)
            except: 
                ticker_upd = None
            price_upd = ticker_upd
    else: 
        # Get the average price 
        price_upd = float(price_upd)/float(steps_ticker - failed_attempts)
        
    return price_upd

##################### Extreme in time series 
# Returns value with the lowest or the highest ticker price among 5-min intervals (candles) 
# type should be 'H' or 'L' (highest ore lowest in the series) 
def candle_extreme(type): 
    global market, candle_steps, candle_sleep
    ticker_upd = {}
    price_upd = 0
    price_extreme = 0
    failed_attempts = 0
    
    for i in range(1, candle_steps + 1): # 5 min: 100 checks x 3 sec (better indication than 30 checks x 10 sec) 
        try:
            #ticker_upd = getticker(exchange, market)
            ticker_upd = coinigy.price(exchange_abbr, market)
            price_upd = ticker_upd
            if type == 'L': 
                if (price_extreme == 0) or (price_upd < price_extreme): 
                    price_extreme = price_upd
            if type == 'H': 
                if (price_extreme == 0) or (price_upd > price_extreme): 
                    price_extreme = price_upd
        except:
            #print "Issues with URL (!) for market", market
            failed_attempts += 1
        time.sleep(candle_sleep) 
        
    # Logging failed attempts number
    if failed_attempts > 0: 
        lprint(["Failed attempts to receive price:", failed_attempts])    
    # If retreiving prices fails completely
    if failed_attempts == steps_ticker:     
        ticker_upd = None # Change
        # Could be related to maintenance
        try:
            send_notification('Maintenance', market + ' seems to be on an automatic maintenance. Will try every 5 minutes.')
        except: 
            lprint(["Failed to send notification"])    
        while ticker_upd is None: 
            time.sleep(300)  
            lprint(["Market could be on maintenance. Sleeping for 5 minutes."])    
            try:
                #ticker_upd = getticker(exchange, market)  
                ticker_upd = coinigy.price(exchange_abbr, market)
            except: 
                ticker_upd = None
            price_upd = ticker_upd
            price_extreme = price_upd
    return price_extreme

##################### Ensuring buy  
def ensure_buy(): 
    global mode 

    if mode == 'now':
        proceed_buy = True  
        lprint(["Proceeding with immediate buy as requested"])
    else: 
        proceed_buy = False  
        price_arr = np.zeros(3)     #3*N-min candlesticks
        lprint(["Running ensure_buy check"])
        
        # Filling the prices array - will be checking for strictly higher lows 
        # this while actually solves the issue of unavailability too
        while (0 in price_arr) or (proceed_buy == False): 
            # Checking Telegram requests and answering 
            approved_flag = check_cancel_flag()
            if approved_flag == False: 
                return 
        
            price_lowest = candle_extreme('L')  
            price_arr = np.append(price_arr, price_lowest)
            price_arr = np.delete(price_arr, [0])

            # Selling on the series of strictly higher lows
            if (0 not in price_arr): 
                lprint(["Low in the candle:", price_lowest, "| higher lows:", strictly_increasing(price_arr)])   
                if strictly_increasing(price_arr): 
                    proceed_buy = True
            else: 
                lprint(["Low in the candle:", price_lowest])   

    return proceed_buy     


##################### Checking balance and changing the position size
def ensure_balance(): 
    global wf_id, job_id, wf_run_mode, bitmex_margin
    global approved_flag  
    global source_position, price_curr
    global comission_rate, exchange

    if wf_run_mode == 's' or wf_run_mode == 'sns':
        return True 
    else: 
        print "Checking balance", exchange, trade
        balance = getbalance(exchange, trade)
        balance_avail = balance['Available']
        
        # Changing the available balance and changing the value if there is no enough funds       
        if exchange == 'bitmex': 
            # rounding to 4 decimals on bitmex 
            #balance_avail = round(Decimal(str(balance_avail)), 5)  # does not work this way
            balance_avail = Decimal(balance_avail).quantize(Decimal('.0001'), rounding='ROUND_DOWN')
            if balance_avail * Decimal(0.99) * bitmex_margin < source_position:         # 0.01 is a commission I reckon 
                source_position = balance_avail * bitmex_margin * Decimal(0.99) 
                lprint(['Corrected the position in ensure balance:', balance_avail ])     
                return True
            else: 
                return True 
        else: 
            if Decimal(str(balance_avail)) < Decimal(str(source_position)): 
                source_position = Decimal(str(balance_avail))  
                lprint(['Corrected the position in ensure balance:', source_position ])       
                return True 
            else: 
                return True 
        
        
###################################################################################
############################## Main workflow #########################################
###################################################################################
### 1. Checking availability, balance 
try: 
    #ticker_upd = getticker(exchange, market)
    ticker_upd = coinigy.price(exchange_abbr, market)
    # Ticker could be failing if there is automatic maintenance - then sleep for a while
    if ticker_upd is None: 
        send_notification('Maintenance', market + ' seems to be on an automatic maintenance. Will try every 5 minutes.')
        while ticker_upd is None: 
            time.sleep(300) # sleeping for 5 minutes and checking again
            #ticker_upd = getticker(exchange, market)  
            ticker_upd = coinigy.price(exchange_abbr, market)
    
    if ticker_upd == 'INVALID_MARKET': 
        lprint(['Error: Invalid market'])    
        send_notification('Error', 'Error: Invalid market to buy')
        logger.close_and_exit()
        
    elif fixed_price != None: 
        # Fat fingers protection if a specific price is requested
        price_check = ticker_upd
        ratio = float(fixed_price)/float(price_check)
        if (ratio >= 1) and (mode == 'reg'): 
            err_msg = 'Error: the requested price is higher than the current. The current price is ' + str(price_check)
            send_notification('Error', err_msg)
            lprint([err_msg])
            logger.close_and_exit()

except urllib2.URLError:
    lprint(['Exchange url unavailable to buy'])    
    send_notification('Error', 'Exchange url unavailable to buy')
    logger.close_and_exit()


### 2. Start timer for price switching and handling simulation modes 
timer_start = time.time()

if mode == 'reg-s': 
    wf_run_mode = 's' # simulating
    mode = 'reg' # setting up a regular mode 
if mode == 'brk-s': 
    wf_run_mode = 's' # simulating
    mode = 'brk' # setting up a regular mode 
if mode == '4h-s': 
    wf_run_mode = 's' # simulating
    mode = '4h' # setting up a regular mode 
if mode == 'fullta-s': 
    wf_run_mode = 's' # simulating
    mode = 'fullta' # setting up a regular mode   
  
# Breakout confirmation price 
if mode == 'brk': 
    breakout_target = fixed_price  

# Reduce the number of checked orders if an immediate buy is requested     
if mode == 'now': 
    orders_check = 3

# Checking 4H data availability 
if mode == '4h' and not td_data_available: 
    print "TD data is unavailable, not possible to start the task"
    logger.close_and_exit()
if mode == 'fullta' and not td_data_available and not td_data_extended_available: 
    print "TD data is unavailable, not possible to start the task"
    logger.close_and_exit()
    
### Inserting in the sqlite db if started fine  
sql_string = "INSERT INTO buys(market, abort_flag, price_fixed, price, source_position, mode, exchange) VALUES ('{}', 0, {}, {}, {}, '{}', '{}')".format(market, int(fixed_price_flag), fixed_price, source_position, mode, exchange)
job_id, rows = query_lastrow_id(sql_string)

### 3. Default values and starting balances / contracts 
source_position = Decimal(str(source_position))
ensure_balance() # added to modify the quantity on the start 

source_start = source_position  
fixed_price_starter = False 
float_price_starter = False 

buy_uuid = None
buy_flag = True 
sum_paid = 0
sum_quantity = 0
approved_flag = True
source_filled = 0
issues_notify = True 

### 4. Main buying loop  
while buy_flag and approved_flag: 
    buy_rate = 0        # price 
    
    try:  
        # Checking existing orders
        if buy_uuid != None: 
            ### 4.1. Get information on the existing orders and cancel them
            lprint(['>>> Cancelling:' , buy_uuid, exchange, market])    
            
            # If we are on bitmex - first we need to get the order info and then to cancel 
            if exchange == 'bitmex': 
                order_info = getorder(exchange, market, buy_uuid)
                #print '\n\nORDER INFO', order_info, '\n\n' #DEBUG
                cancel_stat = cancel(exchange, market, buy_uuid)
                time.sleep(5) 
            else: 
                cancel_stat = cancel(exchange, market, buy_uuid)
                time.sleep(10) # wait for it to be cancelled  - 10 sec
                order_info = getorder(exchange, market, buy_uuid)
            
            ### 4.2. Filled / remaining 
            buy_uuid = None             
            quantity_filled = order_info['Quantity'] - order_info['QuantityRemaining']
            # print ">>>>> Quantity filled {}, order_info['Quantity'] {} , order_info['QuantityRemaining']".format(quantity_filled, order_info['Quantity'], order_info['QuantityRemaining']) #DEBUG 
            
            price_unit = order_info['PricePerUnit']
            price_order = order_info['Price']
            
            if price_unit is None: 
                price_unit = 0
            
            if exchange == 'bitmex': 
                if market == 'USD-BTC': 
                    source_filled = Decimal(str(Decimal(quantity_filled)/Decimal(price_unit)))   
                    sum_paid += Decimal(str(source_filled))   # for price averaging   
                    sum_quantity += quantity_filled
                    str_status = 'Filled: {}'.format(source_filled) 
                else: 
                    source_filled = Decimal(str(Decimal(quantity_filled)*Decimal(price_unit)))  
                    sum_paid += Decimal(str(source_filled))   # for price averaging   
                    sum_quantity += quantity_filled
                    str_status = 'Filled: {}'.format(source_filled) 
            else: 
                source_filled = Decimal(str(price_unit * quantity_filled))
                sum_paid += Decimal(str(source_filled))   
                sum_quantity += quantity_filled  
                str_status = '{} filled {}, {} filled: {}'.format(currency, quantity_filled, trade, source_filled) 
            lprint([str_status])    

        ### 4.3. Timer update
        timer_now = time.time()
        timer_diff = (timer_now - timer_start)/60 # in minutes

        # If out of the requested time period - switching to floating prices       
        if (time_restriction > 0) and (timer_diff > time_restriction) and fixed_price_flag: 
            fixed_price_flag = False 
            orders_check = 5
            comm_string = "Fixed price order could not be filled for " + market + " within required timeframe. Switching to market prices now."
            lprint([comm_string])
            send_notification('Update', comm_string)
            sql_string = "UPDATE buys SET price_fixed = {} WHERE job_id = {}".format(int(fixed_price_flag), job_id)     # updating the DB 
            rows = query(sql_string)

        ### 4.4. Checking the cancellation flag
        approved_flag = check_cancel_flag()
        if approved_flag == False: 
            lprint(["Shutdown was requested via Telegram"])   
            cancel_stat = cancel(exchange, market, buy_uuid)
            time.sleep(5) # wait for it to be cancelled 
            sleep_timer = 0
        
        ### 4.5. Updating how much of source position (e.g. BTC) do we have left and placing a buy order if required
        source_position = ( Decimal(str(source_position)) - Decimal(str(source_filled))  ) * Decimal(str(1 - comission_rate)) 
        lprint(["Updated source position considering commission:", source_position])   
        
        if approved_flag: 
            if fixed_price != 0:
                if (mode == 'reg') or (mode == 'now'):  
                    lprint([exchange, market, ': buying for', source_position, '@', fixed_price])    
                elif mode == 'brk': 
                    lprint([exchange, market, ': buying on breakout target', source_position, '@', breakout_target])    
            else: 
                lprint([exchange, market, ': buying for', source_position, '@ market price'])    
        
        ### 4.6. Get the current price value
        price_curr = get_last_price(market)
        if fixed_price_flag != True:      
            sql_string = "UPDATE buys SET price = {} WHERE job_id = {}".format(price_curr, job_id)    
            rows = query(sql_string)

        # 4H mode     
        if mode == '4h' or mode == 'fullta':
            fixed_price_flag = True # treating as a fixed starter 
                
        ### Price conditions with fixed price for different scenarios
        if (fixed_price_flag) and (fixed_price_starter != True): 
            ##  If we have a fixed price, check if the current is close to the target
            if mode == 'reg':  
                # print "Price current * 0.99: {}, price fixed: {}".format(price_curr * 0.99, fixed_price) # DEBUG #
                # Removed 0.99 because we have ensure_sale and buying for the last price then
                if price_curr <= fixed_price:   
                    # Checking if we should buy or whether it is too early
                    # ensure_balance()  # logic changed
                    fixed_price_starter = ensure_buy()    
                    lprint(["Buy trigger", fixed_price_starter])
                    if fixed_price_starter == True: 
                        fixed_price = get_last_price(market)
                        lprint([exchange, market, ': target price', fixed_price, 'reached and confirmed, start placing buy orders'])    
                    # Otherwise, we will continue in the next loop until we get confirmation on the reversal
                else: 
                    lprint([exchange, market, ': target price', fixed_price, 'not reached. Current:', price_curr])    
                    
            ## Mode: If we are waiting for a breakout  - the logic is simple 
            if mode == 'brk':
                if price_curr >= breakout_target:  
                    lprint([exchange, market, ': breakout price', breakout_target, 'reached confirming.'])    
                    # Checking if we should buy or whether the price is jumping back
                    # ensure_balance()  # logic changed
                    fixed_price_starter = ensure_buy()   
                    lprint(["Buy trigger", fixed_price_starter])
                    if fixed_price_starter == True: 
                        fixed_price = get_last_price(market)
                        lprint([exchange, market, ': breakout price', breakout_target, 'reached and confirmed, start placing buy orders'])    
                else: 
                    lprint([exchange, market, ': breakout price', breakout_target, 'not reached. Current:', price_curr])    
            
            ## Mode: If requested to buy now 
            if mode == 'now':
                fixed_price_starter = True 
            
            ### Mode: 4h based on price action, or larger interval auto-based if fullta is used 
            if mode == '4h' or mode == 'fullta':
                time_hour_update = time.strftime("%H")
                if (time_hour_update <> time_hour): 
                    # Updating the current hour and the TD values 
                    lprint(['Updating the candles price data'])    
                    time_hour = time_hour_update
                    bars = td_info.stats(market, exchange_abbr, td_period, 50000, 5)   
                    
                    # Changing short_flag depending on the direction of the larger time interval if we are in the fullta mode 
                    lprint(['> Extended price action direction:', bars_extended['td_direction'].iloc[-1] ])    
                    if mode == 'fullta': 
                        bars_extended = td_info.stats(market, exchange_abbr, td_period_extended, 100000, 5)   
                        if bars_extended['td_direction'].iloc[-1] == 'down': 
                            short_flag = True 
                        else: 
                            short_flag = False   
                        
                # Different conditions depending on long / short: 
                if not short_flag: # LONGS  
                    check_value = bars['high'].iloc[-2] * (1 + diff_threshold)
                    lprint([ '>', exchange, "TD setup (prev bar):", bars['td_setup'].iloc[-2], "| TD direction (this bar):", bars['td_direction'].iloc[-1], "TD direction (prev bar):", bars['td_direction'].iloc[-2] ])       
                    lprint([ '>', exchange, "Checking condition. Price_curr:", price_curr, "| bar high + threshold:", check_value, "| direction:", bars['td_direction'].iloc[-1] ])       
                    if (bars['td_direction'].iloc[-2] == 'up') and (bars['td_direction'].iloc[-1] == 'up') and (price_curr > check_value):  
                        fixed_price_starter = True 
                    else: 
                        lprint(["> Long buy condition not met"])
                else: #SHORTS 
                    check_value = bars['low'].iloc[-2] * (1 - diff_threshold)
                    lprint([ '>', exchange, "TD setup (prev bar):", bars['td_setup'].iloc[-2], "| TD direction (this bar):", bars['td_direction'].iloc[-1], "TD direction (prev bar):", bars['td_direction'].iloc[-2] ])       
                    lprint([ '>', exchange, "Checking condition. Price_curr:", price_curr, "| bar high - threshold:", check_value, "| direction:", bars['td_direction'].iloc[-1] ])       
                    if (bars['td_direction'].iloc[-2] == 'down') and (bars['td_direction'].iloc[-1] == 'down') and (price_curr < check_value):  
                        fixed_price_starter = True 
                    else: 
                        lprint(["> Short buy condition not met"])
            
      
            
        ### If meeting conditions for fixed price - get the current   
        if fixed_price_flag and fixed_price_starter: 
            fixed_price = get_last_price(market)        
            
        ### Price conditions with floating price 
        if (fixed_price_flag != True) and (float_price_starter != True): 
            lprint(["Starting for the floating (market) price"])
            # Checking for non-4h (time interval-based) cases 
            float_price_starter = ensure_buy()   
            lprint(["Buy trigger", float_price_starter])
                              
        ### 4.7. Checking how much is left to buy and setting the price      
        # print "Source filled:", source_filled, " source start:", source_start     #DEBUG 
        ratio = Decimal(source_filled/source_start)
        ratio = ratio.quantize(Decimal('1.01'))

        if (ratio < 0.96 or ratio == 0) and (approved_flag):           
            if ratio > 0: # DEBUG  
                lprint(['Ratio:', ratio])  
            # If we are using market price (smartbuy)
            if fixed_price_flag != True:      
                # Getting prices if we have not specified a fixed one 
                orderbook = getorderbook(exchange, market)
                if float_price_starter: 
                    lprint(['>> Number of orders used:', orders_check])  
                for elem in orderbook[:orders_check]: 
                    buy_rate += elem['Rate']
                buy_rate = round(buy_rate/orders_check, 8)
            else: 
                # If the price is fixed
                buy_rate = fixed_price
                
            # Updating db 
            sql_string = "UPDATE buys SET price = {} WHERE job_id = {}".format(buy_rate, job_id)    
            rows = query(sql_string)
                       
            ### 4.8. Placing buy order when requirements are met
            if (fixed_price_flag and fixed_price_starter) or ((fixed_price_flag != True) and (float_price_starter)):                
                str_status = 'Used rate: {}'.format(buy_rate)  
                lprint([str_status])    
                quantity = round(Decimal(str(source_position))/Decimal(str(buy_rate)), 6)                
 
                if exchange == 'bitmex': # need to do this in contracts because the api returns contracts and not xbt filled           
                    if market == 'USD-BTC': 
                        quantity = round(Decimal(str(source_position)), 6)
                        buy_rate = round(buy_rate, 0) 
                        contracts = round(quantity * buy_rate)   # margin is already accounted for in the main code     
                        print "Quantity (xbt) {}, buy_rate {}, contracts {}".format(quantity, buy_rate, contracts) # DEBUG 
                    else: # All alts are traded vs btc 
                        quantity = round(Decimal(str(source_position)), 6)
                        buy_rate =  round(buy_rate, 20)    
                        contracts = round(quantity / buy_rate)   # margin is already accounted for in the main code     
                        #print "Quantity (xbt) {}, buy_rate {}, contracts {}".format(quantity, buy_rate, contracts) # DEBUG 
    
                str_status = 'Quantity to buy {}'.format(quantity)  
                lprint([str_status])    
                     
                # Simulation mode
                if wf_run_mode == 's' or wf_run_mode == 'sns' or wf_run_mode == 'reg-s' or wf_run_mode == 'brk-s':
                    buy_flag = False 
                    sleep_timer = 0 
                    lprint(['Bought in simulation'])  
                    sum_quantity = quantity
                    avg_price = buy_rate
                    
                # Real mode    
                else:
                    # Double-checking the quantity after we calculated the actual rate  
                    if quantity > 0.0: 
                        # Bitmex is a bit special 
                        if exchange == 'bitmex':  
                            # Open a long or a short depending on the requested side 
                            if short_flag: 
                                print 'Contracts (short) {} buy_rate {}'.format(contracts, buy_rate) #DEBUG    
                                buy_result = selllimit(exchange, market, None, buy_rate, contracts)  
                            else: 
                                print 'Contracts (long) {} buy_rate {}'.format(contracts, buy_rate) #DEBUG    
                                buy_result = buylimit(exchange, market, None, buy_rate, contracts)  
                        else: 
                            print 'Quantity {} buy_rate {}'.format(quantity, buy_rate) #DEBUG    
                            buy_result = buylimit(exchange, market, quantity, buy_rate)  

                        #print "\n>>> Result", buy_result #DEBUG
                        lprint(["-------------------------------------------------------------------- \n>> Result:", buy_result, "\n--------------------------------------------------------------------"])  # DEBUG # 
                        
                        if buy_result == 'MIN_TRADE_REQUIREMENT_NOT_MET': 
                            # If trade requirement were not met or an error occured       
                            buy_flag = False 
                            sleep_timer = 0 
                            chat.send('Cancelling buying on ' + market + ' as minimum trade requirements were not met')
                        else:   
                            # If the requirements were met 
                            try: 
                                buy_uuid = buy_result['uuid']
                                lprint(['>> Placed order', buy_uuid])    
                            except: 
                                # If something else is wrong    
                                buy_flag = False 
                                sleep_timer = 0 
                                err_msg = traceback.format_exc()
                                comm_string = '{} issue. Reason: {}'.format(market, err_msg)
                                lprint([comm_string])    
                                chat.send(comm_string) 
                                # Also cancelling workflow until we figure out what's up 
                                wf_id = None
                    else: 
                        buy_flag = False 
                        sleep_timer = 0 
            
            # Number of orders to check 
            if orders_check > 1: 
                orders_check -= 1 # 5,4,3,2,1
            
        # If we got all we need     
        else: 
            buy_flag = False 
            sleep_timer = 0 
        
        # Sleeping depending on whether we have started buying or not 
        if ratio == 0: 
            time.sleep(30)
        else: 
            time.sleep(sleep_timer)

    except: 
        err_msg = traceback.format_exc()
        comm_string = '{} buy task stopped. {} {} spent. Reason: {}'.format(market, sum_paid, trade, err_msg)
        lprint([comm_string])    
        send_notification('Error', comm_string)
        # Cancel the orders 
        if buy_uuid != None: 
            # Get info and cancel 
            lprint(['Cancelling:' , buy_uuid])    
            cancel_stat = cancel(exchange, market, buy_uuid)
        # Deleting the task from the db 
        sql_string = "DELETE FROM buys WHERE job_id = {}".format(job_id)
        rows = query(sql_string)

        # Delete workflow for this market
        if wf_id is not None: 
            sql_string = "DELETE FROM workflow WHERE wf_id = {}".format(wf_id)
            rows = query(sql_string)
            wf_id = None
        logger.close_and_exit()
 

### 5. Deleting the task from the db 
sql_string = "DELETE FROM buys WHERE job_id = {}".format(job_id)
rows = query(sql_string)

'''
#DEBUG block  
print ">>>>Afterbuy WF ID", wf_id
print "Sum_q", sum_quantity
print "Sum_paid", sum_paid
print "Average price paid", round(sum_paid / sum_quantity, 8)
''' 
### 6. Calculating averages and updating the information / logging the results 
if sum_quantity > 0: 
    if (wf_run_mode != 's') and (wf_run_mode != 'sns'):
        if exchange == 'bitmex': 
            if market == 'USD-BTC': 
                avg_price = round(Decimal(sum_quantity) / Decimal(str(sum_paid)), 8)    # cause we are buying contracts there  
            else: 
                avg_price = round(Decimal(sum_paid) / Decimal(str(sum_quantity)), 8)    # cause we are buying contracts there  
        else: 
            avg_price = round(Decimal(sum_paid) / Decimal(str(sum_quantity)), 8)
        lprint(['Average price paid:', avg_price])    
    else:
        # If simulation
        sum_paid = source_position
    
    # Description
    if short_flag: 
        direction_desc = 'short'
    else: 
        direction_desc = 'long'
    
    if exchange == 'bitmex': 
        comm_string = "{}: buy orders completed on {}, opened a position for {} contracts. Direction: {}".format(market, exchange, sum_quantity, direction_desc)
    else: 
        comm_string = "{}: buy orders completed on {} at the total amount of {} and the average price {}.\nBought {} {}.".format(market, exchange, sum_paid, avg_price, sum_quantity, currency)
    send_notification('Bought', comm_string)
    lprint([comm_string])    
    
    # Updating workflow info if we have a workflow 
    if wf_id is not None: 
        sql_string = "UPDATE workflow SET sum_q = '{}', avg_price = '{}' WHERE wf_id = {}".format(sum_quantity, avg_price, wf_id) 
        job_id, rows = query_lastrow_id(sql_string)

    # Logging the results
    if (wf_run_mode != 's') and (wf_run_mode != 'sns'):
        try: 
            date_time = strftime("%Y-%m-%d %H:%M", localtime())
            wb = load_workbook(config.trade_hist_filename)
            ws = wb['Entry_points']
            if exchange == 'bitmex': 
                new_line = [date_time, 'XBT', avg_price, sum_quantity, sum_paid]
            else: 
                new_line = [date_time, currency, avg_price, sum_quantity, sum_paid]
            ws.append(new_line)
            max_row = ws.max_row
            # Apply a style 
            index_row = "{}:{}".format(max_row, max_row) 
            for cell in ws[index_row]:
                cell.font = Font(name='Arial', size=10)
            #
            wb.save(config.trade_hist_filename)
            ''' 
            # Used locally 
            if platform_run != 'Windows': 
                copyfile('/home/illi4/Robot/Trade_history.xlsx', '/mnt/hgfs/Shared_folder/Trade_history.xlsx')
            ''' 
        except: 
            lprint(['Trade history xls unavailable']) 

else: 
    send_notification('Cancelled', exchange + ':' + market + ': buy order was cancelled and nothing was bought')
    lprint([market, ': buy order was cancelled and nothing was bought'])    
    # Delete workflow for this market
    if wf_id is not None: 
        sql_string = "DELETE FROM workflow WHERE wf_id = {}".format(wf_id)
        rows = query(sql_string)
        wf_id = None

# Closing logs properly but not exiting     
logger.close()
        
### 7. If this is a workflow - launching a new task 
if wf_id is not None: 
    sql_string = "SELECT * FROM workflow WHERE wf_id = '{}'".format(wf_id)
    rows = query(sql_string)

    try: 
        wf_info = rows[0]   # first result if existing 
        wf_info_market = wf_info[1]
        wf_info_trade = wf_info[2]
        wf_info_curr = wf_info[3]
        wf_info_price = wf_info[8]
        wf_stop_mode = wf_info[9]
        wf_info_tp = wf_info[4]
        wf_info_sl = wf_info[5]
        wf_price_entry = wf_info[10]
        exchange_abbr = wf_info[11]
        
        # If buyback 
        if wf_price_entry is not None: 
            if float(wf_price_entry) > 0: 
                wf_info_price = wf_price_entry
        
        # Deleting wf_id from the db 
        sql_string = "DELETE FROM workflow WHERE wf_id = {}".format(wf_id)
        rows = query(sql_string)
        
        # The short_flag can be inconsistent with the original record if the fullta mode (automatic) was used and the direction flipped - handling this here
        if (not short_flag and (float(wf_info_tp) < float(wf_info_sl))) or (short_flag and (float(wf_info_tp) > float(wf_info_sl)) ):
            lprint(['Flipping TP and SL values in line with the direction. Short_flag:' , short_flag])    
            wf_info_tp, wf_info_sl = wf_info_sl, wf_info_tp
                 
        print '>>> Start a profit task: {} {} {} {} {} {}'.format(wf_stop_mode, exchange_abbr, wf_info_market, wf_info_price, wf_info_tp, wf_info_sl)

        # Launch in the same window 
        python_call = 'python robot.py ' + ' '.join([wf_stop_mode, exchange_abbr, wf_info_market, str(wf_info_price), str(wf_info_tp), str(wf_info_sl)]) 
        p = subprocess.Popen(python_call, shell=True, stderr=subprocess.PIPE)
        while True:
            out = p.stderr.read(1)
            if out == '' and p.poll() != None:
                break
            if out != '':
                sys.stdout.write(out)
                sys.stdout.flush()
                
    except:
        chat.send('Could not launch sell task in the workflow for ' + market + ' on ' + exchange + ' or the task finished with an error')





